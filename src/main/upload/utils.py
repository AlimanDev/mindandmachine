import pandas as pd
import re
import time
import datetime
import paramiko
import json
from src.util.utils import JsonResponse
from src.conf.djconfig import ALLOWED_UPLOAD_EXTENSIONS
from django.utils.datastructures import MultiValueDictKeyError
from django.conf import settings
from functools import wraps
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from dateutil.relativedelta import relativedelta
from src.db.models import (
    User,
    Shop,
    PeriodClients,
    WorkType,
    WorkerDay,
    WorkerDayCashboxDetails,
    OperationType,
    WorkerCashboxInfo,
    WorkerPosition,
    AttendanceRecords,
    Employment,
)
from src.main.demand.utils import create_predbills_request_function
from src.util.models_converter import Converter
from src.conf.djconfig import SFTP_IP, SFTP_PASSWORD, SFTP_USERNAME, SFTP_PATH

WORK_TYPES = {
    'В': WorkerDay.TYPE_HOLIDAY,
    'ОТ': WorkerDay.TYPE_VACATION,
}


def get_uploaded_file(func):
    """
    Проверят загруженный на сервак файл(есть ли вообще файл в запросе и какого он формата)
    18.11.2018 -- пока поддерживаем только excel

    запускать с api_method
    Args:
        request(WSGIrequest): request

    Returns:
        file
    """
    @wraps(func)
    def wrapper(request, form, *args, **kwargs):
        try:
            file = request.FILES['file']
        except MultiValueDictKeyError:
            return JsonResponse.value_error('Не было передано ни одного файла.')

        if not file:
            return JsonResponse.value_error('Файл не был загружен.')
        if not file.name.split('/')[-1].split('.', file.name.split('/')[-1].count('.'))[-1] in ALLOWED_UPLOAD_EXTENSIONS:
            return JsonResponse.value_error('Файлы с таким расширением не поддерживается.')

        return func(request, form, file, *args, **kwargs)
    return wrapper

def upload_vacation_util(vacation_file):
    """
    Принимает от клиента экселевский файл и загружает отпуска

    Args:
         method: POST
    """

    try:
        worksheet = pd.read_csv(vacation_file, sep=';')
    except KeyError:
        return JsonResponse.internal_error('Не удалось открыть активный лист.')

    tabel_code_col = 0
    from_dt_col = 1
    to_dt_col = 2

    worksheet[from_dt_col] = pd.to_datetime(worksheet.iloc[:, from_dt_col], format='%d.%m.%y')
    worksheet[to_dt_col] = pd.to_datetime(worksheet.iloc[:, to_dt_col], format='%d.%m.%y')

    list_to_create = []
    for index, row in worksheet.iterrows():
        try:
            user = User.objects.get(tabel_code=row[tabel_code_col])
        except Exception:
            return JsonResponse.internal_error('Не удалось найти пользователя {}'.format(row[tabel_code_col]))

        from_dt = row[from_dt_col].date()
        to_dt = row[to_dt_col].date()

        if to_dt < from_dt:
            return JsonResponse.internal_error('Дата конца отпуска должна быть больше или равна дате начала отпуска.')

        while from_dt <= to_dt:
            wd_query_set = list(WorkerDay.objects.filter(dt=from_dt, worker=user).order_by('-id'))
            if len(wd_query_set):
                WorkerDayCashboxDetails.objects.filter(
                    worker_day__in=wd_query_set
                ).delete()
                for wd in wd_query_set:
                    wd.delete()

            if len(list_to_create) >= 999:
                WorkerDay.objects.bulk_create(list_to_create)
                list_to_create = []

            list_to_create.append(
                WorkerDay(
                    worker=user,
                    type=WorkerDay.TYPE_VACATION,
                    dt=from_dt,
                    dttm_work_start=None,
                    dttm_work_end=None,
                )
            )
            from_dt += datetime.timedelta(days=1)

    WorkerDay.objects.bulk_create(list_to_create)

    return JsonResponse.success()

def upload_timetable_util(form, timetable_file):
    """
    Принимает от клиента экселевский файл и создает расписание (на месяц)

    Args:
         method: POST
         url: /api/upload/upload_timetable
         shop_id(int): required = True
    """
    shop_id = form['shop_id']
    shop = Shop.objects.get(id=shop_id)

    try:
        worksheet = load_workbook(timetable_file).active
    except KeyError:
        return JsonResponse.internal_error('Не удалось открыть активный лист.')

    ######################### сюда писать логику чтения из экселя ######################################################
    fio_column = 2
    work_type_column = 3
    workers_start_row = 19
    dates_row = 17
    dates_start_column = 5

    work_dates = []
    for column in worksheet.iter_cols(min_col=dates_start_column, min_row=dates_row, max_row=dates_row):
        for cell in column:
            if isinstance(cell.value, datetime.datetime):
                work_dates.append(cell.value.date())
    dates_end_column = dates_start_column + len(work_dates) - 1
    if not work_dates:
        return JsonResponse.value_error('Не смог сгенерировать массив дат. Возможно они в формате строки.')
    shop_work_types = {w.name: w for w in WorkType.objects.filter(shop_id=shop_id, dttm_deleted__isnull=True)}
    for row in worksheet.iter_rows(min_row=workers_start_row):
        user_work_type = None
        for cell in row:
            column_index = column_index_from_string(cell.column)
            if column_index == fio_column:
                first_last_names = cell.value.split(' ')
                last_name_concated = ' '.join(first_last_names[:-2])
                username = str(time.time() * 1000000)[:-2]
                # try:
                u, create = User.objects.get_or_create(
                    shop_id=shop_id,
                    last_name=last_name_concated,
                    first_name=first_last_names[0],
                    defaults={
                        'username': username,
                    }
                )
                # except User.DoesNotExist:
                #     return JsonResponse.value_error('Не могу найти пользователя на строке {}'.format(cell.row))
                if create:
                    employment = Employment.objects.create(
                        user=u,
                        shop_id=shop_id,
                    )
                    u.username = 'u' + str(u.id)
                    u.save()
                else:
                    employment, _ = Employment.objects.get_or_create(
                        user=u,
                        shop_id=shop_id,
                    )
            elif column_index == work_type_column:
                user_work_type = shop_work_types.get(cell.value, None)
                if user_work_type:
                    WorkerCashboxInfo.objects.get_or_create(
                        employment=employment,
                        work_type=user_work_type,
                    )

            if dates_start_column <= column_index <= dates_end_column:
                dt = work_dates[column_index - dates_start_column]
                dttm_work_start = None
                dttm_work_end = None
                #  todo: если будут типы с цифрами, надо будет переделать
                if bool(re.search(r'\d', cell.value)):
                    times = cell.value.split('-')
                    work_type = WorkerDay.TYPE_WORKDAY
                    dttm_work_start = datetime.datetime.combine(
                        dt, Converter.parse_time(times[0] + ':00')
                    )
                    dttm_work_end = datetime.datetime.combine(
                        dt, Converter.parse_time(times[1] + ':00')
                    )
                    if dttm_work_end < dttm_work_start:
                        dttm_work_end += datetime.timedelta(days=1)
                else:
                    work_type = WORK_TYPES[cell.value]

                wd_query_set = WorkerDay.objects.filter(dt=dt, worker=u)
                WorkerDayCashboxDetails.objects.filter(
                    worker_day__in=wd_query_set
                ).delete()
                for wd in wd_query_set.order_by('-id'):  # потому что могут быть родители у wd
                    wd.delete()
                work_hours = 0
                if (work_type in WorkerDay.TYPES_PAID):
                    break_triplets = json.loads(shop.break_triplets)
                    work_hours = WorkerDay.count_work_hours(break_triplets, dttm_work_start, dttm_work_end)
                new_wd = WorkerDay.objects.create(
                    worker=u,
                    employment=employment,
                    shop_id=shop_id,
                    dt=dt,
                    work_hours=work_hours,
                    dttm_work_start=dttm_work_start,
                    dttm_work_end=dttm_work_end,
                    type=work_type
                )
                if work_type == WorkerDay.TYPE_WORKDAY:
                    WorkerDayCashboxDetails.objects.create(
                        worker_day=new_wd,
                        dttm_from=dttm_work_start,
                        dttm_to=dttm_work_end,
                        status=WorkerDayCashboxDetails.TYPE_WORK,
                        work_type=user_work_type,
                    )

    ####################################################################################################################

    return JsonResponse.success()

def upload_demand_util(demand_file, form=None):
    code_col = 0
    datetime_col = 1
    value_col = 2

    try:
        worksheet = pd.read_csv(demand_file, sep=';')
    except KeyError:
        return JsonResponse.internal_error('Не удалось открыть активный лист.')

    worksheet[datetime_col] = pd.to_datetime(worksheet.iloc[:, datetime_col])
    worksheet.iloc[:, datetime_col] = worksheet[datetime_col].map(lambda x: x.replace(minute=0, second=0))
    from_dttm = worksheet.iloc[:, datetime_col].min()
    to_dttm = worksheet.iloc[:, datetime_col].max()

    if demand_file.name.split('/')[-1] == 'incoming.csv':
        shop_codes = worksheet.iloc[:,code_col].unique()
        name_operation_type = 'intro'
    else:
        # Нет определённого пользователся (но может он из магазина, который у нас будет)
        users = User.objects.filter(tabel_code__in=worksheet.iloc[:, code_col].unique()).select_related('shop')
        shop_tabel = {user.tabel_code:user.shop.super_shop.code for user in users}
        shop_codes = {value for value in shop_tabel.values()}
        worksheet = worksheet.groupby([
            worksheet.columns[code_col],
            worksheet.columns[datetime_col],
        ])[worksheet.columns[value_col]].sum().reset_index()
        name_operation_type = 'bills'

    for shop_code in shop_codes:
        list_to_create = []
        try:
            operation_type = OperationType.objects.get(
                work_type__shop__super_shop__code=shop_code,
                name=name_operation_type,
            )
        except Exception:
            print('Нет такого operation_type: {} - {}'.format(shop_code, name_operation_type))
            continue

        PeriodClients.objects.filter(
            dttm_forecast__range=[from_dttm, to_dttm],
            type=PeriodClients.FACT_TYPE,
            operation_type=operation_type,
        ).delete()

        for index, row in worksheet.iterrows():
            if demand_file.name.split('/')[-1] == 'incoming.csv' and row[code_col] == shop_code or \
                   demand_file.name.split('/')[-1] != 'incoming.csv' and shop_tabel[str(row[code_col])] == shop_code:
                if len(list_to_create) == 999:
                    PeriodClients.objects.bulk_create(list_to_create)
                    list_to_create = []
                list_to_create.append(PeriodClients(
                    dttm_forecast=row[datetime_col],
                    type=PeriodClients.FACT_TYPE,
                    operation_type=operation_type,
                    value=row[value_col],
                ))

        PeriodClients.objects.bulk_create(list_to_create)

        from_dt_to_create = PeriodClients.objects.filter(
            type=PeriodClients.FACT_TYPE,
            operation_type__name=operation_type.name,
        ).order_by('dttm_forecast').last().dttm_forecast.date() + relativedelta(days=1)

        result_of_func = create_predbills_request_function(shop_id=operation_type.work_type.shop.id, dt=from_dt_to_create)

    return JsonResponse.success() if result_of_func is True else result_of_func


def upload_employees_util(vacation_file):
    """
    Принимает от клиента экселевский файл и загружает отпуска

    Args:
         method: POST
    """

    try:
        worksheet = pd.read_csv(vacation_file, sep=';')
    except KeyError:
        return JsonResponse.internal_error('Не удалось открыть активный лист.')

    tabel_code_col = 0
    last_name_col = 1
    first_name_col = 2
    middle_name_col = 3
    position_col = 4
    hired_col = 5
    fired_col = 6
    shop_code_col = 7

    worksheet[hired_col] = pd.to_datetime(worksheet.iloc[:, hired_col], format='%d.%m.%y')
    worksheet[fired_col] = pd.to_datetime(worksheet.iloc[:, fired_col], format='%d.%m.%y')

    for index, row in worksheet.iterrows():
        update_kwargs = {
            'tabel_code': row[tabel_code_col],
            'last_name': row[last_name_col],
            'first_name': row[first_name_col],
            'middle_name': row[middle_name_col],
            'position': WorkerPosition.objects.get(title=row[position_col]),
            'dt_hired': row[hired_col],
            'dt_fired': row[fired_col],
            'shop': Shop.objects.filter(super_shop__code=row[shop_code_col])[0],
        }
        user, create = User.objects.update_or_create(
            tabel_code=row[tabel_code_col],
            last_name=row[last_name_col],
            first_name=row[first_name_col],
            middle_name=row[middle_name_col],
            defaults=update_kwargs,
        )

        if create:
            user.username = 'user-' + str(user.id)
            user.save()

    return JsonResponse.success()


def upload_urv_util(urv_file):

    try:
        worksheet = pd.read_excel(urv_file)
    except KeyError:
        return JsonResponse.internal_error('Не удалось открыть активный лист.')

    col_fio = 4
    col_date = 5
    col_coming = 6
    col_leaving = 7
    dt_format = '%Y-%m-%d'
    dttm_format = '%Y-%m-%d %H:%M:%S'

    from_dt = datetime.datetime.strptime(worksheet.iloc[:, col_date].min().split(' ')[0], dt_format)
    to_dt = datetime.datetime.strptime(worksheet.iloc[:, col_date].max().split(' ')[0], dt_format) + datetime.timedelta(days=1)

    list_to_create = []
    user_list = []
    not_found_user = ''
    for index, row in worksheet.iterrows():
        if row[col_fio] != 'NULL' and str(row[col_fio]) != 'nan':
            fio = row[col_fio].split(' ')
            if fio != not_found_user:
                current_dt = row[col_date].split(' ')[0]

                try:
                    user = User.objects.get(first_name=fio[1], last_name=fio[0], dt_fired__isnull=True)
                    user_list.append(user)
                except Exception:
                    not_found_user = fio
                    print('Пользователь {} {} {} не найден или уволен.'.format(fio[0], fio[1], fio[2]))
                    continue

                if len(list_to_create) >= 999:
                    AttendanceRecords.objects.bulk_create(list_to_create)
                    list_to_create = []

                if row[col_coming] != 'NULL' and str(row[col_coming]) != 'nan':
                    dttm_coming = datetime.datetime.strptime(
                        '{} {}'.format(current_dt, row[col_coming].split('.')[0]), dttm_format)

                    list_to_create.append(
                        AttendanceRecords(
                            type=AttendanceRecords.TYPE_COMING,
                            dttm=dttm_coming,
                            user=user,
                            shop=user.shop,
                        ))
                if row[col_leaving] != 'NULL' and str(row[col_leaving]) != 'nan':
                    dttm_leaving = datetime.datetime.strptime(
                        '{} {}'.format(current_dt, row[col_leaving].split('.')[0]), dttm_format)

                    list_to_create.append(
                        AttendanceRecords(
                            type=AttendanceRecords.TYPE_LEAVING,
                            dttm=dttm_leaving,
                            user=user,
                            shop=user.shop,
                        ))
    AttendanceRecords.objects.filter(dttm__range=[from_dt, to_dt], user__in=user_list).delete()
    AttendanceRecords.objects.bulk_create(list_to_create)
    return JsonResponse.success()


def sftp_download(localpath):
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(SFTP_IP, username=SFTP_USERNAME, password=SFTP_PASSWORD)
    sftp = ssh.open_sftp()
    sftp.get(SFTP_PATH + localpath, localpath)
    sftp.close()
    ssh.close()
