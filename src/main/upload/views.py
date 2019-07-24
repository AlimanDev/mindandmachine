import datetime

from openpyxl import load_workbook
import pandas as pd
import re
import time
from openpyxl.utils import column_index_from_string
from dateutil.relativedelta import relativedelta
from src.util.utils import api_method, JsonResponse
from .utils import get_uploaded_file, WORK_TYPES
from .forms import UploadForm
from src.db.models import (
    User,
    PeriodClients,
    WorkType,
    Notifications,
    WorkerDay,
    WorkerDayCashboxDetails,
    OperationType,
    WorkerCashboxInfo,
)
from src.main.demand.utils import create_predbills_request_function
from src.util.models_converter import BaseConverter


@api_method('POST', UploadForm)
@get_uploaded_file
def upload_demand(request, form, demand_file):
    """
    Принимает от клиента экселевский файл в формате из TPNET (для леруа специально) и загружает из него данные в бд

    Args:
         method: POST
         url: /api/upload/upload_demand
         shop_id(int): required = True
    """
    try:
        worksheet = load_workbook(demand_file).active
    except KeyError:
        return JsonResponse.internal_error('Не удалось открыть активный лист.')

    list_to_create = []

    shop_id = form['shop_id']
    cash_column_num = 0
    time_column_num = 1
    value_column_num = 2
    datetime_format = '%d.%m.%Y %H:%M:%S'
    processed_rows = 0  # счетчик "обработынных" строк
    checked_rows = 0  # счетчик "проверенных" строк
    from_to = []

    def create_demand_objs(obj=None):
        if obj is None:
            PeriodClients.objects.bulk_create(list_to_create)
        elif len(list_to_create) == 999:
            list_to_create.append(obj)
            PeriodClients.objects.bulk_create(list_to_create)
            list_to_create[:] = []
        else:
            list_to_create.append(obj)

    # work_types = list(WorkType.objects.filter(shop_id=shop_id))
    operation_types = list(OperationType.objects.filter(work_type__shop_id=shop_id).select_related('work_type'))
    # todo: check conflicts in operation_types_names or user other format
    operation_types = {op_type.name or op_type.work_type.name: op_type for op_type in operation_types}
    ######################### сюда писать логику чтения из экселя ######################################################
    for index, row in enumerate(worksheet.rows):
        dttm = row[time_column_num].value
        if '.' not in dttm or ':' not in dttm:
            continue
        if not isinstance(dttm, datetime.datetime):
            try:
                dttm = datetime.datetime.strptime(dttm, datetime_format)
                from_to.append(dttm)
            except ValueError:
                return JsonResponse.value_error(
                    'Невозможно преобразовать время. Пожалуйста введите формат {}'.format(datetime_format))

    from_to.sort()
    dttm_from = from_to[1]
    dttm_to = from_to[-1]

    PeriodClients.objects.filter(
        dttm_forecast__range=[dttm_from, dttm_to],
        type=PeriodClients.FACT_TYPE,
        operation_type__work_type__shop_id=shop_id,
    ).delete()

    for index, row in enumerate(worksheet.rows):
        value = row[value_column_num].value
        if isinstance(row[value_column_num].value, str):
            #  может быть 'Нет данных'
            continue
        checked_rows += 1
        dttm = row[time_column_num].value
        if '.' not in dttm or ':' not in dttm:
            continue
        if not isinstance(dttm, datetime.datetime):
            try:
                dttm = datetime.datetime.strptime(dttm, datetime_format)
            except ValueError:
                return JsonResponse.value_error(
                    'Невозможно преобразовать время. Пожалуйста введите формат {}'.format(datetime_format))

        cashtype_name = row[cash_column_num].value
        cashtype_name = cashtype_name[:1].upper() + cashtype_name[1:].lower()  # учет регистра чтобы нормально в бд было
        ct_to_search = operation_types.get(cashtype_name, None)

        if ct_to_search:
            create_demand_objs(
                PeriodClients(
                    dttm_forecast=dttm,
                    operation_type=ct_to_search,
                    value=value,
                    type=PeriodClients.FACT_TYPE
                )
            )
            processed_rows += 1

        else:
            return JsonResponse.internal_error('Невозможно прочитать тип работ на строке №{}.'.format(index))
    if processed_rows == 0 or processed_rows < checked_rows / 2:  # если было обработано, меньше половины строк, че-то пошло не так
        return JsonResponse.internal_error(
            'Было обработано {}/{} строк. Пожалуйста, проверьте формат данных в файле.'.format(
                processed_rows, worksheet.max_row
            )
        )

    ####################################################################################################################

    create_demand_objs(None)

    # works only for postgres
    # unique_fact = list(PeriodClients.objects.filter(
    #     type=PeriodClients.FACT_TYPE,
    #     operation_type__work_type__shop_id=shop_id
    # ).order_by('dttm_forecast', 'operation_type_id', '-id').distinct('dttm_forecast', 'operation_type_id'))
    #
    # PeriodClients.objects.filter(
    #     type=PeriodClients.FACT_TYPE,
    #     operation_type__work_type__shop_id=shop_id
    # ).delete()
    # PeriodClients.objects.bulk_create(unique_fact)
    from_dt_to_create = PeriodClients.objects.filter(
        type=PeriodClients.FACT_TYPE,
        operation_type__work_type__shop_id=shop_id
    ).order_by('dttm_forecast').last().dttm_forecast.date() + relativedelta(days=1)

    result_of_func = create_predbills_request_function(shop_id=shop_id, dt=from_dt_to_create)

    return JsonResponse.success() if result_of_func is True else result_of_func


@api_method('POST', UploadForm)
@get_uploaded_file
def upload_timetable(request, form, timetable_file):
    """
    Принимает от клиента экселевский файл и создает расписание (на месяц)

    Args:
         method: POST
         url: /api/upload/upload_timetable
         shop_id(int): required = True
    """
    shop_id = form['shop_id']

    try:
        worksheet = load_workbook(timetable_file).active
    except KeyError:
        return JsonResponse.internal_error('Не удалось открыть активный лист.')

    ######################### сюда писать логику чтения из экселя ######################################################
    fio_column = 2
    work_type_column = 3
    workers_start_row = 19
    dates_row = 17
    dates_start_column = 5

    work_dates = []
    for column in worksheet.iter_cols(min_col=dates_start_column, min_row=dates_row, max_row=dates_row):
        for cell in column:
            if isinstance(cell.value, datetime.datetime):
                work_dates.append(cell.value.date())
    dates_end_column = dates_start_column + len(work_dates) - 1
    if not work_dates:
        return JsonResponse.value_error('Не смог сгенерировать массив дат. Возможно они в формате строки.')
    shop_work_types = {w.name: w for w in WorkType.objects.filter(shop_id=shop_id, dttm_deleted__isnull=True)}
    for row in worksheet.iter_rows(min_row=workers_start_row):
        user_work_type = None
        for cell in row:
            column_index = column_index_from_string(cell.column)
            if column_index == fio_column:
                first_last_names = cell.value.split(' ')
                last_name_concated = ' '.join(first_last_names[:-2])
                username = str(time.time() * 1000000)[:-2]
                # try:
                u, create = User.objects.get_or_create(
                    shop_id=shop_id,
                    last_name=last_name_concated,
                    first_name=first_last_names[0],
                    defaults={
                        'username': username,
                    }
                )
                # except User.DoesNotExist:
                #     return JsonResponse.value_error('Не могу найти пользователя на строке {}'.format(cell.row))
                if create:
                    u.username = 'u' + str(u.id)
                    u.save()
            elif column_index == work_type_column:
                user_work_type = shop_work_types.get(cell.value, None)
                if user_work_type:
                    WorkerCashboxInfo.objects.get_or_create(
                        worker=u,
                        work_type=user_work_type,
                    )

            if dates_start_column <= column_index <= dates_end_column:
                dt = work_dates[column_index - dates_start_column]
                dttm_work_start = None
                dttm_work_end = None
                #  todo: если будут типы с цифрами, надо будет переделать
                if bool(re.search(r'\d', cell.value)):
                    times = cell.value.split('-')
                    work_type = WorkerDay.Type.TYPE_WORKDAY.value
                    dttm_work_start = datetime.datetime.combine(
                        dt, BaseConverter.parse_time(times[0] + ':00')
                    )
                    dttm_work_end = datetime.datetime.combine(
                        dt, BaseConverter.parse_time(times[1] + ':00')
                    )
                    if dttm_work_end < dttm_work_start:
                        dttm_work_end += datetime.timedelta(days=1)
                else:
                    work_type = WORK_TYPES[cell.value]

                wd_query_set = WorkerDay.objects.filter(dt=dt, worker=u)
                WorkerDayCashboxDetails.objects.filter(
                    worker_day__in=wd_query_set
                ).delete()
                for wd in wd_query_set.order_by('-id'):  # потому что могут быть родители у wd
                    wd.delete()
                new_wd = WorkerDay.objects.create(
                    worker=u,
                    dt=dt,
                    dttm_work_start=dttm_work_start,
                    dttm_work_end=dttm_work_end,
                    type=work_type
                )
                if work_type == WorkerDay.Type.TYPE_WORKDAY.value:
                    WorkerDayCashboxDetails.objects.create(
                        worker_day=new_wd,
                        dttm_from=dttm_work_start,
                        dttm_to=dttm_work_end,
                        status=WorkerDayCashboxDetails.TYPE_WORK,
                        work_type=user_work_type,
                    )

    ####################################################################################################################

    return JsonResponse.success()
