import json
import re
from datetime import datetime

from django.db import transaction
from openpyxl import load_workbook

from src.base.models import Shop, Network

LOAD_SHOP_SCHEDULE_PATTERN = re.compile(
    r'(?P<when>пн-вс|будни|выходные|пн-пт|сб-вс|вс-чт|пт-сб|сб|вс):?(?P<open>\d{1,2}[:.]\d{2})-(?P<close>\d{1,2}[:.]\d{2})')

MONDAY = 0
TUESDAY = 1
WEDNESDAY = 2
THURSDAY = 3
FRIDAY = 4
SATURDAY = 5
SUNDAY = 6

HOLIDAYS = [
    SATURDAY,
    SUNDAY,
]
WEEKDAYS = [
    MONDAY,
    TUESDAY,
    WEDNESDAY,
    THURSDAY,
    FRIDAY,
]
# ALL = WEEKDAYS + HOLIDAYS

days_mapping = {
    'пн-вс': '__all__',
    'будни': WEEKDAYS,
    'выходные': HOLIDAYS,
    'пн-пт': WEEKDAYS,
    'сб-вс': HOLIDAYS,
    'вс-чт': [
        SUNDAY,
        MONDAY,
        TUESDAY,
        WEDNESDAY,
        THURSDAY,
    ],
    'пт-сб': [
        FRIDAY,
        SATURDAY,
    ],
    'сб': [
        SATURDAY,
    ],
    'вс': [
        SUNDAY,
    ]
}


def _prepare_time(time_string):
    try:
        time = datetime.strptime(time_string, "%H:%M")
    except ValueError:
        time = datetime.strptime(time_string, "%H.%M")
    except:
        raise
    return time.strftime('%H:%M:%S')


def load_schedule(xlsx_path, logs_path, start_row, end_row, cols, network_name='Ортека'):
    with transaction.atomic():
        with open(logs_path, 'w') as f:
            network = Network.objects.get(name=network_name)

            ws = load_workbook(xlsx_path).active
            for row in ws.iter_rows(min_row=start_row, max_col=cols, max_row=end_row):
                shop_status, shop_code, schedule_string = list(map(lambda i: i.value, row))

                try:
                    shop = Shop.objects.get(code=shop_code, network=network)
                except Shop.DoesNotExist:
                    print(f'no shop with code={shop_code}', file=f)
                    continue

                schedule_string = re.sub(r"\s+", "", schedule_string.lower())

                schedules = re.findall(LOAD_SHOP_SCHEDULE_PATTERN, schedule_string)
                if not schedules:
                    print(f'can\'t parse schedule_string={schedule_string}', file=f)
                    continue

                tm_open_dict = {}
                tm_close_dict = {}
                for pattern, tm_open, tm_close in schedules:
                    days = days_mapping.get(pattern)
                    if not days:
                        print(f'can\'t find days for pattern={pattern}', file=f)
                        continue

                    if days == '__all__':
                        tm_open_dict['all'] = _prepare_time(tm_open)
                        tm_close_dict['all'] = _prepare_time(tm_close)
                    else:
                        for day in days:
                            tm_open_dict[day] = _prepare_time(tm_open)
                            tm_close_dict[day] = _prepare_time(tm_close)

                shop.tm_open_dict = json.dumps(tm_open_dict)
                shop.tm_close_dict = json.dumps(tm_close_dict)
                shop.save(update_fields=('tm_open_dict', 'tm_close_dict'))
